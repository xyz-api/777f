# 通用锁班助手

import re
import platform
import os
from datetime import datetime
from colorama import init, Fore, Style
from playwright.sync_api import sync_playwright

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

init()  # 初始化colorama


def beep_error():
    """错误提示音"""
    try:
        if platform.system() == 'Windows':
            import winsound
            winsound.Beep(800, 300)
        else:
            os.system('paplay /usr/share/sounds/freedesktop/stereo/dialog-error.oga 2>/dev/null || true')
    except:
        pass

# 锁班类型映射表 (显示名 -> 代码)
LEAVE_TYPE_MAP = {
    # 假期类
    "ALV-年假（公休假）": "ALV",
    "ALV_FD-飞行员公休（订座）": "ALV_FD",
    "RECU_LVE-健康疗养": "RECU_LVE",
    "RECU_LVE_R-康复疗养": "RECU_LVE_R",
    "MAT_FA_LVE-陪产假": "MAT_FA_LVE",
    "MAT_MO_LVE-产假": "MAT_MO_LVE",
    "PREGNANT-孕假": "PREGNANT",
    "PARENT_LVE-探亲假-探父母": "PARENT_LVE",
    "SPOUSE_LVE-探亲假-探配偶": "SPOUSE_LVE",
    "MARR_LVE-婚假": "MARR_LVE",
    "COMP_LVE-丧假": "COMP_LVE",
    "CHILD_LVE-育儿假": "CHILD_LVE",
    "INJURY_LVE-工伤假": "INJURY_LVE",
    "LWOP_LVE-其他（事假）": "LWOP_LVE",
    "UNPAID_LVE-无薪": "UNPAID_LVE",
    "HOUSE_LVE-搬家": "HOUSE_LVE",
    "BREED_LVE-哺乳假": "BREED_LVE",
    "PATERNITY-独生子女护理假": "PATERNITY",
    "BIRC_LVE-计划生育假": "BIRC_LVE",
    "REWARD_LVE-奖励": "REWARD_LVE",
    "PENALTY-停飞": "PENALTY",
    "PRD_LVE-经期假": "PRD_LVE",
    # 值勤期类
    "GRD-地面班": "GRD",
    "GDO-地面休息": "GDO",
    "TRNG1-训练": "TRNG1",
    "BS_STUDY-业务学习": "BS_STUDY",
    "BUSINESS-公务": "BUSINESS",
    "GRD_ONDUTY-地面值班": "GRD_ONDUTY",
    "LG_STUDY-语言学习/考试": "LG_STUDY",
    "MEDL_CHK-体检_临床": "MEDL_CHK",
    "MEDL_PHLE-体检_抽血": "MEDL_PHLE",
    "MEDL_EET-体检_平板": "MEDL_EET",
    "MEDL_PSYC-体检_心理测试": "MEDL_PSYC",
    "MTG-会议": "MTG",
    "MTG_SF-安全讲评会": "MTG_SF",
    "DGET-危险品培训": "DGET",
    "EP-飞行人员应急复训": "EP",
    "CRM-CRM培训": "CRM",
    "T_SIM_INS-模拟机检查": "T_SIM_INS",
    "T_SIM_REC-模拟机复训": "T_SIM_REC",
    "T_SIM_INT-模拟机初始": "T_SIM_INT",
    "T_SIM_UPG-模拟机升级": "T_SIM_UPG",
    "T_SIM_CON-模拟机_转机型": "T_SIM_CON",
    "MAKEUP-补考": "MAKEUP",
    "BS_CONCL-飞行后讲评": "BS_CONCL",
    "BS_CHK-业务检查": "BS_CHK",
    "ADMN-管理任务": "ADMN",
    "SOCIAL-社会活动": "SOCIAL",
    "HANDBOOK-手册": "HANDBOOK",
    "POL_STUDY-政治学习": "POL_STUDY",
    "T/A-部门活动": "T/A",
}

# 代码到中文名的反向映射
LEAVE_CODE_TO_NAME = {v: k for k, v in LEAVE_TYPE_MAP.items()}


def parse_leave_type(text: str) -> str:
    """解析锁班类型，支持代码或中文名"""
    if not text:
        return None
    text = str(text).strip()
    # 直接是代码
    if text in LEAVE_CODE_TO_NAME:
        return text
    # 完整格式 "CODE-中文名"
    if text in LEAVE_TYPE_MAP:
        return LEAVE_TYPE_MAP[text]
    # 只有中文名，模糊匹配
    for key, code in LEAVE_TYPE_MAP.items():
        name_part = key.split('-', 1)[1] if '-' in key else key
        if name_part in text or text in name_part:
            return code
    return None


def parse_excel_file(filepath: str, whitelist: set = None) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    if not HAS_OPENPYXL:
        return [], ["未安装openpyxl库，请运行: pip install openpyxl"]
    
    records = []
    errors = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # 跳过表头，从第2行开始
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            # 列顺序: 员工号、姓名、锁班类型、开始日期、结束日期
            emp_id = str(row[0]).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            leave_type_raw = str(row[2]).strip() if len(row) > 2 and row[2] else None
            start_date = row[3] if len(row) > 3 else None
            end_date = row[4] if len(row) > 4 else None
            
            # 验证员工号
            if not emp_id or not re.match(r'^\d{6}$', emp_id):
                if emp_id and emp_id != 'None':
                    errors.append(f"第{row_num}行: 员工号格式错误 [{emp_id}]")
                continue
            
            # 白名单过滤
            if whitelist and emp_id not in whitelist:
                continue
            
            # 解析锁班类型
            leave_type = parse_leave_type(leave_type_raw)
            if not leave_type:
                errors.append(f"第{row_num}行: 未识别锁班类型 [{leave_type_raw}]")
                continue
            
            # 解析日期
            def format_date(d):
                if isinstance(d, datetime):
                    return d.strftime('%Y-%m-%d')
                if isinstance(d, str):
                    return normalize_date(d)
                return None
            
            start = format_date(start_date)
            end = format_date(end_date) if end_date else start
            
            if not start:
                errors.append(f"第{row_num}行: 日期格式错误")
                continue
            
            records.append({
                "员工号": emp_id,
                "姓名": name,
                "请假类型": leave_type,
                "开始日期": start,
                "结束日期": end or start
            })
        
        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")
    
    return records, errors


def c_info(text):
    return f"{Fore.CYAN}{text}{Style.RESET_ALL}"

def c_ok(text):
    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"

def c_err(text):
    return f"{Fore.RED}{text}{Style.RESET_ALL}"

def c_warn(text):
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"

def c_hint(text):
    return f"{Fore.MAGENTA}{text}{Style.RESET_ALL}"


def parse_whitelist(text: str) -> set:
    """解析员工号白名单"""
    all_nums = re.findall(r'\d{6}', re.sub(r'\D', ' ', text))
    if all_nums:
        return set(all_nums)
    text = re.sub(r'\D', '', text)
    return set(text[i:i+6] for i in range(0, len(text), 6) if len(text[i:i+6]) == 6)


def normalize_date(date_str: str) -> str:
    """把各种日期格式统一转成YYYY-MM-DD"""
    parts = re.split(r'[-/]', date_str)
    if len(parts) == 3:
        year, month, day = parts
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return date_str


def parse_single_record(text: str) -> dict:
    """解析单条记录"""
    result = {"员工号": None, "姓名": None, "请假类型": None, "开始日期": None, "结束日期": None}
    emp = re.search(r'\b(\d{6})\b', text)
    if emp:
        result["员工号"] = emp.group(1)
    name = re.search(r'\d{6}\s*([\u4e00-\u9fa5]{2,4})', text)
    if name:
        result["姓名"] = name.group(1)
    # 使用新的解析函数
    for key, val in LEAVE_TYPE_MAP.items():
        if key in text or val in text:
            result["请假类型"] = val
            break
    dates = re.findall(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', text)
    if dates:
        result["开始日期"] = normalize_date(dates[0])
        result["结束日期"] = normalize_date(dates[1]) if len(dates) > 1 else normalize_date(dates[0])
    return result


def split_continuous_text(text: str) -> list:
    """把连续粘贴的文本按员工号切分成多条记录"""
    # 按6位员工号切分
    parts = re.split(r'(?=\d{6}[\u4e00-\u9fa5])', text)
    return [p.strip() for p in parts if p.strip() and re.search(r'\d{6}', p)]


def parse_batch_input(text: str, whitelist: set = None) -> tuple:
    """解析批量输入"""
    records = []
    errors = []
    # 先按换行分，如果只有一行且很长，尝试按员工号切分
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])
    for i, line in enumerate(lines, 1):
        record = parse_single_record(line)
        if whitelist and record["员工号"] and record["员工号"] not in whitelist:
            continue
        if not record["员工号"]:
            errors.append(f"第{i}条: 未识别员工号 [{line[:50]}]")
            continue
        if not record["请假类型"]:
            errors.append(f"第{i}条: 未识别请假类型 [{line[:50]}]")
            continue
        if not record["开始日期"]:
            errors.append(f"第{i}条: 未识别日期 [{line[:50]}]")
            continue
        records.append(record)
    return records, errors


def clear_form(page):
    """清空表单"""
    page.locator("#showIdshowNonproductionTaskImportPage").fill("")
    page.locator("#lockStartTime").fill("")
    page.locator("#lockEndTime").fill("")


def fill_form(page, emp_id, leave_type, start_date, end_date):
    """填写表单"""
    clear_form(page)
    emp_input = page.locator("#showIdshowNonproductionTaskImportPage")
    emp_input.click()
    emp_input.fill("")
    emp_input.type(str(emp_id), delay=10)
    page.evaluate("""
        const input = document.querySelector('#showIdshowNonproductionTaskImportPage');
        if (input) {
            input.dispatchEvent(new Event('blur', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
        }
    """)
    page.wait_for_timeout(1000)
    # 用JS直接设置下拉框值并触发事件
    page.evaluate("""(leaveType) => {
        const select = document.querySelector('#lockType');
        if (select) {
            select.focus();
            select.value = leaveType;
            select.dispatchEvent(new MouseEvent('mousedown', { bubbles: true }));
            select.dispatchEvent(new MouseEvent('click', { bubbles: true }));
            select.dispatchEvent(new Event('input', { bubbles: true }));
            select.dispatchEvent(new Event('change', { bubbles: true }));
            select.dispatchEvent(new Event('blur', { bubbles: true }));
        }
    }""", leave_type)
    page.wait_for_timeout(500)
    page.locator("#lockStartTime").fill(start_date)
    page.locator("#lockEndTime").fill(end_date)


def submit_and_check(page):
    """提交并检查冲突,返回(成功, 冲突信息)"""
    # 点击下一步
    page.get_by_role("button", name="下一步").wait_for()
    page.get_by_role("button", name="下一步").click()
    # 等待继续录入按钮出现,说明页面加载完成
    page.get_by_role("button", name="继续录入").wait_for()
    # 检查查询结果是否有数据
    result_rows = page.locator("#showNonproductionTaskImportResultPage1 tbody.list tr")
    # 检查冲突列表的内容
    conflict_rows = page.locator("#showNonproductionTaskImportResultPage2 tbody.list tr")
    # 获取冲突列表的文本内容
    conflict_text = ""
    if conflict_rows.count() > 0:
        conflict_text = conflict_rows.first.inner_text()
    # 成功条件: 查询结果有数据 且 冲突列表显示"没有相关信息"
    if result_rows.count() > 0 and "没有相关信息" in conflict_text:
        # 没有冲突,点击继续录入
        page.get_by_role("button", name="继续录入").click()
        # 等待表单页面加载
        page.locator("#showIdshowNonproductionTaskImportPage").wait_for()
        return True, None
    else:
        # 有冲突或查询结果为空
        if result_rows.count() == 0:
            conflict_info = "查询结果为空"
        else:
            conflict_info = conflict_text
        return False, conflict_info


def whitelist_status(whitelist):
    """返回白名单状态文字"""
    if whitelist:
        return c_ok(f"白名单:{len(whitelist)}人")
    return c_warn("白名单:无")


def read_multiline(prompt, confirm_key='ok', cancel_key='c'):
    """读取多行输入,输入confirm_key确认,cancel_key取消"""
    print(prompt)
    lines = []
    while True:
        line = input()
        if line.lower() == cancel_key:
            return None
        if line.lower() == confirm_key:
            break
        if line:
            lines.append(line)
    if not lines:
        return None
    return '\n'.join(lines)


def set_whitelist():
    """设置白名单"""
    text = read_multiline(c_hint("请粘贴员工号列表(输入ok确认,c取消):"), 'ok', 'c')
    if text is None:
        print(c_warn("已取消"))
        return None
    wl = parse_whitelist(text)
    if not wl:
        print(c_err("未识别到有效员工号"))
        return None
    print(c_ok(f"已设置白名单,共{len(wl)}人"))
    return wl


def format_record(r):
    """格式化记录显示"""
    name = r['姓名'] or '未知'
    return f"{r['员工号']} {name} {r['请假类型']} {r['开始日期']}~{r['结束日期']}"


def go_back_to_form(page):
    """从结果页返回表单页"""
    try:
        page.get_by_role("button", name="继续录入").click()
        page.locator("#showIdshowNonproductionTaskImportPage").wait_for()
    except Exception:
        pass  # 如果已经在表单页就忽略


def print_failed_records(failed_records):
    """打印失败记录并写入日志文件"""
    if failed_records:
        print(c_err(f"本次失败{len(failed_records)}条:"))
        for r, reason in failed_records:
            print(c_err(f"  {format_record(r)} - {reason}"))
        # 写入日志文件
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"failed_{timestamp}.txt"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(f"失败记录 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"共{len(failed_records)}条\n")
                f.write("-" * 50 + "\n")
                for r, reason in failed_records:
                    f.write(f"{format_record(r)} - {reason}\n")
            filepath = os.path.abspath(filename)
            print(c_warn(f"失败记录已保存: {filepath}"))
        except Exception as e:
            print(c_warn(f"保存日志失败: {e}"))


def batch_mode(page, whitelist):
    """批量模式"""
    failed_records = []  # 记录失败的条目
    while True:
        print(f"{c_info('[批量模式]')} {whitelist_status(whitelist)}")
        text = read_multiline(c_hint("请粘贴数据(输入ok确认,b返回):"), 'ok', 'b')
        if text is None:
            print_failed_records(failed_records)
            return
        records, errors = parse_batch_input(text, whitelist)
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(err))
        if not records:
            print(c_err("没有可处理的记录"))
            continue
        print(c_ok(f"共{len(records)}条有效数据:"))
        for i, r in enumerate(records, 1):
            print(f"{i}. {format_record(r)}")
        confirm = input(c_hint("y开始填写,n重新粘贴,b返回主菜单: ")).strip().lower()
        if confirm == 'b':
            return
        if confirm != 'y':
            continue
        i = 0
        while i < len(records):
            record = records[i]
            print(f"{c_info(f'[{i+1}/{len(records)}]')} 填写: {format_record(record)}")
            try:
                fill_form(page, record["员工号"], record["请假类型"], record["开始日期"], record["结束日期"])
                print(c_ok("填表完成,提交中..."))
                success, conflict_info = submit_and_check(page)
                if success:
                    print(c_ok("提交成功"))
                    i += 1
                else:
                    beep_error()
                    print(c_err("有冲突!"))
                    print(c_warn(conflict_info if conflict_info else "未知冲突"))
                    failed_records.append((record, "有冲突"))
                    go_back_to_form(page)
                    i += 1
            except Exception as e:
                beep_error()
                print(c_err(f"失败: {e}"))
                failed_records.append((record, str(e)))
                i += 1
        print(c_ok("批量处理完成"))
        print_failed_records(failed_records)
        return


def manual_mode(page, whitelist):
    """手动模式"""
    while True:
        print(f"{c_info('[手动模式]')} {whitelist_status(whitelist)} | {c_hint('粘贴数据,b返回主菜单:')}")
        text = input().strip()
        if text.lower() == 'b':
            return
        if not text:
            continue
        record = parse_single_record(text)
        if whitelist and record["员工号"] and record["员工号"] not in whitelist:
            print(c_err("该员工不在白名单中"))
            continue
        if not record["员工号"]:
            print(c_err("未识别员工号"))
            continue
        if not record["请假类型"]:
            print(c_err("未识别请假类型"))
            continue
        if not record["开始日期"]:
            print(c_err("未识别日期"))
            continue
        while True:
            print(f"填写: {format_record(record)}")
            try:
                fill_form(page, record["员工号"], record["请假类型"], record["开始日期"], record["结束日期"])
                print(c_ok("填表完成,提交中..."))
                success, conflict_info = submit_and_check(page)
                if success:
                    print(c_ok("提交成功"))
                    break
                else:
                    print(c_err("有冲突!"))
                    print(c_warn(conflict_info if conflict_info else "未知冲突"))
                    go_back_to_form(page)
                    break
            except Exception as e:
                beep_error()
                print(c_err(f"失败: {e}"))
                break


def excel_mode(page, whitelist):
    """Excel导入模式"""
    if not HAS_OPENPYXL:
        print(c_err("未安装openpyxl库，请运行: pip install openpyxl"))
        return
    
    failed_records = []
    while True:
        print(f"{c_info('[Excel导入]')} {whitelist_status(whitelist)}")
        filepath = input(c_hint("请输入Excel文件路径(b返回): ")).strip()
        if filepath.lower() == 'b':
            print_failed_records(failed_records)
            return
        
        # 去除引号
        filepath = filepath.strip('"').strip("'")
        
        if not os.path.exists(filepath):
            print(c_err("文件不存在"))
            continue
        
        records, errors = parse_excel_file(filepath, whitelist)
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(f"  {err}"))
        
        if not records:
            print(c_err("没有可处理的记录"))
            continue
        
        print(c_ok(f"共{len(records)}条有效数据:"))
        for i, r in enumerate(records, 1):
            print(f"{i}. {format_record(r)}")
        
        confirm = input(c_hint("y开始填写,n重新选择,b返回主菜单: ")).strip().lower()
        if confirm == 'b':
            return
        if confirm != 'y':
            continue
        
        i = 0
        while i < len(records):
            record = records[i]
            print(f"{c_info(f'[{i+1}/{len(records)}]')} 填写: {format_record(record)}")
            try:
                fill_form(page, record["员工号"], record["请假类型"], record["开始日期"], record["结束日期"])
                print(c_ok("填表完成,提交中..."))
                success, conflict_info = submit_and_check(page)
                if success:
                    print(c_ok("提交成功"))
                    i += 1
                else:
                    beep_error()
                    print(c_err("有冲突!"))
                    print(c_warn(conflict_info if conflict_info else "未知冲突"))
                    failed_records.append((record, "有冲突"))
                    go_back_to_form(page)
                    i += 1
            except Exception as e:
                beep_error()
                print(c_err(f"失败: {e}"))
                failed_records.append((record, str(e)))
                i += 1
        print(c_ok("Excel导入完成"))
        print_failed_records(failed_records)
        return


def main():
    print(c_info("通用锁班助手"))
    # 浏览器路径
    browser_path = input(c_hint("浏览器路径(回车用默认): ")).strip() or None
    if browser_path:
        print(c_ok(f"使用指定浏览器: {browser_path}"))
    else:
        print(c_ok("使用默认浏览器"))
    # 白名单
    whitelist = None
    use_wl = input(c_hint("是否预设白名单?(y/n): ")).strip().lower()
    if use_wl == 'y':
        whitelist = set_whitelist()
    else:
        print(c_ok("不设置白名单,处理所有员工"))
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=False, executable_path=browser_path)
    context = browser.new_context()
    context.set_default_timeout(0)  # 全局无超时限制
    page = context.new_page()
    # 登录
    try:
        page.goto("https://ieb.csair.com/login")
        page.wait_for_load_state("networkidle")
        page.locator("#scanLogin").wait_for()
        page.locator("#scanLogin").click()
        print(c_info("请扫码登录..."))
        page.wait_for_url("**/index/**")
        page.wait_for_load_state("networkidle")
        print(c_ok("登录成功"))
    except Exception as e:
        print(c_err(f"自动登录失败: {e}"))
        print(c_warn("请手动完成登录"))
        input(c_hint("登录完成后按回车继续..."))
    # 导航到非生产任务录入页面
    try:
        print(c_info("正在进入非生产任务录入页面..."))
        page.goto("https://ieb.csair.com/index/index")
        page.wait_for_load_state("networkidle")
        page.get_by_text("运行管理").nth(1).wait_for()
        page.get_by_text("运行管理").nth(1).click()
        page.get_by_role("link", name="非生产任务").wait_for()
        page.get_by_role("link", name="非生产任务").click()
        page.get_by_role("link", name="非生产任务录入").wait_for()
        page.get_by_role("link", name="非生产任务录入").click()
        page.locator("#mainContent").wait_for()
        page.locator("#mainContent").click()
        page.wait_for_load_state("networkidle")
        print(c_ok("已进入非生产任务录入页面"))
    except Exception as e:
        print(c_err(f"自动导航失败: {e}"))
        print(c_warn("请手动进入非生产任务录入页面"))
        input(c_hint("准备好后按回车继续..."))
    print(c_ok("开始工作"))
    while True:
        print(f"{whitelist_status(whitelist)} | {c_hint('1批量 2手动 3Excel导入 w设白名单 c清白名单 q退出')}")
        cmd = input(c_hint("选择: ")).strip().lower()
        if cmd == '1':
            batch_mode(page, whitelist)
        elif cmd == '2':
            manual_mode(page, whitelist)
        elif cmd == '3':
            excel_mode(page, whitelist)
        elif cmd == 'w':
            new_wl = set_whitelist()
            if new_wl is not None:
                whitelist = new_wl
        elif cmd == 'c':
            whitelist = None
            print(c_ok("已清除白名单"))
        elif cmd == 'q':
            break
    browser.close()
    pw.stop()
    print(c_info("结束"))


if __name__ == "__main__":
    main()
