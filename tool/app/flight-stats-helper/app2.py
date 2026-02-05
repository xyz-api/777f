# 飞行经历起落数查询助手 - 批量模式（统信浏览器版本）

import re
import os
from datetime import datetime
from colorama import init, Fore, Style
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

init()


def c_info(text):
    return f"{Fore.CYAN}{text}{Style.RESET_ALL}"

def c_ok(text):
    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"

def c_err(text):
    return f"{Fore.RED}{text}{Style.RESET_ALL}"

def c_warn(text):
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"

def c_hint(text):
    return f"{Fore.MAGENTA}{text}{Style.RESET_ALL}"


def normalize_date(date_str: str) -> str:
    """把各种日期格式统一转成YYYY/MM/DD"""
    parts = re.split(r'[-/]', str(date_str))
    if len(parts) == 3:
        year, month, day = parts
        return f"{year}/{month.zfill(2)}/{day.zfill(2)}"
    return date_str


def parse_single_record(text: str) -> dict:
    """解析单条记录"""
    result = {"员工号": None, "姓名": None, "开始日期": None, "结束日期": None}
    
    # 提取员工号
    emp = re.search(r'\b(\d{6})\b', text)
    if emp:
        result["员工号"] = emp.group(1)
    
    # 提取姓名（紧跟员工号后的中文）
    name = re.search(r'\d{6}\s*([\u4e00-\u9fa5]{2,4})', text)
    if name:
        result["姓名"] = name.group(1)
    
    # 提取日期
    dates = re.findall(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', text)
    if dates:
        result["开始日期"] = normalize_date(dates[0])
        result["结束日期"] = normalize_date(dates[1]) if len(dates) > 1 else normalize_date(dates[0])
    
    return result


def split_continuous_text(text: str) -> list:
    """把连续粘贴的文本按员工号切分成多条记录"""
    parts = re.split(r'(?=\d{6}[\u4e00-\u9fa5])', text)
    return [p.strip() for p in parts if p.strip() and re.search(r'\d{6}', p)]


def parse_batch_input(text: str) -> tuple:
    """解析批量输入"""
    records = []
    errors = []
    
    # 先按换行分，如果只有一行且很长，尝试按员工号切分
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])
    
    for i, line in enumerate(lines, 1):
        record = parse_single_record(line)
        
        if not record["员工号"]:
            errors.append(f"第{i}条: 未识别员工号 [{line[:50]}]")
            continue
        
        if not record["开始日期"]:
            errors.append(f"第{i}条: 未识别日期 [{line[:50]}]")
            continue
        
        records.append(record)
    
    return records, errors


def read_multiline(prompt, confirm_key='ok', cancel_key='c'):
    """读取多行输入"""
    print(prompt)
    lines = []
    while True:
        line = input()
        if line.lower() == cancel_key:
            return None
        if line.lower() == confirm_key:
            break
        if line:
            lines.append(line)
    if not lines:
        return None
    return '\n'.join(lines)


def parse_excel_file(filepath: str) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    records = []
    errors = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            emp_id = str(int(row[0])).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            start_date = row[2] if len(row) > 2 else None
            end_date = row[3] if len(row) > 3 else None
            
            if not emp_id or not re.match(r'^\d{6}$', emp_id):
                if emp_id and emp_id != 'None':
                    errors.append(f"第{row_num}行: 员工号格式错误 [{emp_id}]")
                continue
            
            def format_date(d):
                if isinstance(d, datetime):
                    return d.strftime('%Y/%m/%d')
                if isinstance(d, str):
                    return normalize_date(d)
                return None
            
            start = format_date(start_date)
            end = format_date(end_date) if end_date else start
            
            if not start:
                errors.append(f"第{row_num}行: 日期格式错误")
                continue
            
            records.append({
                "员工号": emp_id,
                "姓名": name,
                "开始日期": start,
                "结束日期": end or start,
                "行号": row_num
            })
        
        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")
    
    return records, errors


def write_to_excel(filepath: str, row_num: int, flight_exp: str, landing_count: str):
    """实时写入Excel"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 写入飞行经历（E列）和起落总数（F列）
        ws.cell(row=row_num, column=5, value=flight_exp)
        ws.cell(row=row_num, column=6, value=landing_count)
        
        wb.save(filepath)
        wb.close()
        return True
    except Exception as e:
        print(c_err(f"写入Excel失败: {e}"))
        return False


def fill_date(page, date_str):
    """填写日期到日期选择器 - 统信浏览器版本"""
    parts = date_str.split('/')
    if len(parts) != 3:
        raise ValueError(f"日期格式错误: {date_str}")
    
    year, month, day = parts
    
    # 月份映射 - 注意11月和12月的特殊格式
    month_map = {
        '01': '一月', '02': '二月', '03': '三月', '04': '四月',
        '05': '五月', '06': '六月', '07': '七月', '08': '八月',
        '09': '九月', '10': '十月', '11': '十一', '12': '十二'
    }
    
    month_cn = month_map.get(month, month)
    day_str = str(int(day))  # 去掉前导0
    
    # 等待iframe加载
    page.wait_for_timeout(300)
    
    # 使用frame_locator代替locator().content_frame
    frame = page.frame_locator("iframe >> nth=2")
    
    # 1. 点击年份输入框
    frame.get_by_role("textbox").nth(1).click()
    page.wait_for_timeout(100)
    
    # 2. 选择年份
    frame.get_by_role("cell", name=year).click()
    page.wait_for_timeout(100)
    
    # 3. 点击月份输入框
    frame.get_by_role("textbox").first.click()
    page.wait_for_timeout(100)
    
    # 4. 选择月份
    frame.get_by_role("cell", name=month_cn).click()
    page.wait_for_timeout(100)
    
    # 5. 选择日期 - 使用first选择第一个匹配的日期（避免匹配到15、25等）
    frame.get_by_role("cell", name=day_str).first.click()
    page.wait_for_timeout(100)


def query_flight_record(page, emp_id, start_date, end_date, clear_first=False):
    """执行查询操作"""
    # 填写员工号
    emp_input = page.get_by_placeholder("员工号或姓名")
    emp_input.click()
    page.wait_for_timeout(100)
    
    if clear_first:
        # 统信浏览器：先清空再输入
        emp_input.fill("")
        page.wait_for_timeout(100)
    
    # 用type模拟逐字输入
    emp_input.type(str(emp_id), delay=50)
    page.wait_for_timeout(300)
    
    # 填写开始日期 - 多次点击确保触发日期选择器
    page.locator("#flyTimeExperience_beginDate").click()
    page.wait_for_timeout(200)
    if clear_first:
        # 第二次及以后需要多点几次
        page.locator("#flyTimeExperience_beginDate").click()
        page.wait_for_timeout(200)
    fill_date(page, start_date)
    
    # 填写结束日期 - 多次点击确保触发日期选择器
    page.locator("#flyTimeExperience_endDate").click()
    page.wait_for_timeout(200)
    if clear_first:
        # 第二次及以后需要多点几次
        page.locator("#flyTimeExperience_endDate").click()
        page.wait_for_timeout(200)
    fill_date(page, end_date)
    
    # 点击查询
    page.wait_for_timeout(300)
    page.get_by_role("button", name="查询").click()
    page.wait_for_timeout(1500)


def extract_flight_data(page):
    """提取飞行经历数据"""
    try:
        # 等待表格加载
        page.wait_for_timeout(1000)
        
        # 查找表格tbody
        tbody = page.locator("tbody.list")
        
        # 获取第一行数据
        first_row = tbody.locator("tr").first
        
        # 提取所有td单元格
        cells = first_row.locator("td")
        cell_count = cells.count()
        
        if cell_count == 0:
            return None
        
        # 提取飞行经历（第9列）和起落总数（第16列）
        data = {
            "飞行经历": cells.nth(8).inner_text().strip() if cell_count > 8 else "",
            "起落总数": cells.nth(15).inner_text().strip() if cell_count > 15 else "",
        }
        
        return data
    except Exception as e:
        print(c_warn(f"提取数据失败: {e}"))
        return None


def main():
    print(c_info("飞行经历起落数查询助手 - 批量模式（统信浏览器版本）"))
    
    # 启动浏览器
    print(c_info("正在启动浏览器..."))
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=False)
    context = browser.new_context()
    context.set_default_timeout(0)
    page = context.new_page()
    
    try:
        # 登录
        page.goto("https://ieb.csair.com/login")
        page.wait_for_load_state("networkidle")
        page.locator("#scanLogin").wait_for()
        page.locator("#scanLogin").click()
        print(c_info("请扫码登录..."))
        page.wait_for_url("**/index/**")
        page.wait_for_load_state("networkidle")
        print(c_ok("登录成功"))
    except Exception as e:
        print(c_err(f"自动登录失败: {e}"))
        print(c_warn("请手动完成登录"))
        input(c_hint("登录完成后按回车继续..."))
    
    try:
        # 导航到飞行经历页面 - 统信浏览器版本
        print(c_info("正在进入飞行经历查询页面..."))
        page.goto("https://ieb.csair.com/index/index")
        page.wait_for_load_state("networkidle")
        # 统信浏览器的统计应用点击方式
        page.get_by_role("listitem").filter(has_text="统计应用").locator("span").click()
        page.get_by_role("link", name="综合报表").wait_for()
        page.get_by_role("link", name="综合报表").click()
        page.get_by_role("link", name="飞行经历").wait_for()
        page.get_by_role("link", name="飞行经历").click()
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(500)
        page.get_by_role("radio").nth(2).check()
        page.wait_for_timeout(300)
        print(c_ok("已进入飞行经历查询页面"))
    except Exception as e:
        print(c_err(f"自动导航失败: {e}"))
        print(c_warn("请手动进入飞行经历查询页面"))
        input(c_hint("准备好后按回车继续..."))
    
    # 选择输入方式
    print(c_hint("\n请选择输入方式:"))
    print("1. 读取Excel文件")
    print("2. 直接粘贴数据")
    choice = input(c_hint("请选择(1/2): ")).strip()
    
    records = []
    output_file = None
    
    if choice == '1':
        # Excel文件模式
        filepath = input(c_hint("请输入Excel文件路径: ")).strip().strip('"').strip("'")
        
        if not os.path.exists(filepath):
            print(c_err("文件不存在"))
            return
        
        print(c_info("正在读取Excel文件..."))
        records, errors = parse_excel_file(filepath)
        output_file = filepath
        
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(f"  {err}"))
    
    elif choice == '2':
        # 粘贴数据模式
        text = read_multiline(c_hint("请粘贴数据(输入ok确认,c取消):"), 'ok', 'c')
        
        if text is None:
            print(c_warn("已取消"))
            return
        
        print(c_info("正在解析数据..."))
        records, errors = parse_batch_input(text)
        
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(err))
        
        # 创建新的Excel文件
        if records:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"飞行经历查询结果_{timestamp}.xlsx"
            
            # 创建Excel并写入表头和数据
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["员工号", "姓名", "起始时间", "截止时间", "飞行经历", "起落总数"])
            
            for i, record in enumerate(records, start=2):
                ws.cell(row=i, column=1, value=record["员工号"])
                ws.cell(row=i, column=2, value=record.get("姓名", ""))
                ws.cell(row=i, column=3, value=record["开始日期"])
                ws.cell(row=i, column=4, value=record["结束日期"])
                record["行号"] = i  # 记录行号
            
            wb.save(output_file)
            wb.close()
            print(c_ok(f"已创建输出文件: {output_file}"))
    
    else:
        print(c_err("无效选择"))
        return
    
    if not records:
        print(c_err("没有可处理的记录"))
        return
    
    print(c_ok(f"共{len(records)}条有效数据"))
    for i, r in enumerate(records, 1):
        name = r.get('姓名', '未知')
        print(f"{i}. {r['员工号']} {name} {r['开始日期']}~{r['结束日期']}")
    
    # 确认开始
    confirm = input(c_hint("按回车开始查询，输入q退出: ")).strip().lower()
    if confirm == 'q':
        return
    
    # 批量查询
    print(c_ok("开始批量查询"))
    success_count = 0
    fail_count = 0
    
    for i, record in enumerate(records):
        name = record.get('姓名', '未知')
        print(f"{c_info(f'[{i+1}/{len(records)}]')} 查询: {record['员工号']} {name} {record['开始日期']}~{record['结束日期']}")
        
        try:
            # 查询
            query_flight_record(page, record["员工号"], record["开始日期"], record["结束日期"], clear_first=(i > 0))
            
            # 提取数据
            data = extract_flight_data(page)
            
            if data and data.get('飞行经历') and data.get('起落总数'):
                flight_exp = data['飞行经历']
                landing_count = data['起落总数']
                
                # 实时写入Excel
                if write_to_excel(output_file, record['行号'], flight_exp, landing_count):
                    print(c_ok(f"✓ 飞行经历: {flight_exp} | 起落总数: {landing_count} | 已写入Excel"))
                    success_count += 1
                else:
                    print(c_err(f"✗ 查询成功但写入失败"))
                    fail_count += 1
            else:
                print(c_warn(f"✗ 未查询到数据"))
                write_to_excel(output_file, record['行号'], "无数据", "无数据")
                fail_count += 1
                
        except Exception as e:
            print(c_err(f"✗ 失败: {e}"))
            write_to_excel(output_file, record['行号'], "查询失败", "查询失败")
            fail_count += 1
    
    # 完成
    print(c_ok(f"\n批量查询完成！成功: {success_count}, 失败: {fail_count}"))
    print(c_ok(f"结果已保存到: {os.path.abspath(output_file)}"))
    print(c_info("\n浏览器保持打开状态，可以手动查看结果"))
    print(c_hint("按回车关闭程序..."))
    input()


if __name__ == "__main__":
    main()
