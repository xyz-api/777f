# 飞行经历起落数查询助手

import re
import platform
import os
from datetime import datetime
from colorama import init, Fore, Style
from playwright.sync_api import sync_playwright

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

init()


def beep_error():
    """错误提示音"""
    try:
        if platform.system() == 'Windows':
            import winsound
            winsound.Beep(800, 300)
        else:
            os.system('paplay /usr/share/sounds/freedesktop/stereo/dialog-error.oga 2>/dev/null || true')
    except:
        pass


def c_info(text):
    return f"{Fore.CYAN}{text}{Style.RESET_ALL}"

def c_ok(text):
    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"

def c_err(text):
    return f"{Fore.RED}{text}{Style.RESET_ALL}"

def c_warn(text):
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"

def c_hint(text):
    return f"{Fore.MAGENTA}{text}{Style.RESET_ALL}"


def parse_whitelist(text: str) -> set:
    """解析员工号白名单"""
    all_nums = re.findall(r'\d{6}', re.sub(r'\D', ' ', text))
    if all_nums:
        return set(all_nums)
    text = re.sub(r'\D', '', text)
    return set(text[i:i+6] for i in range(0, len(text), 6) if len(text[i:i+6]) == 6)


def normalize_date(date_str: str) -> str:
    """把各种日期格式统一转成YYYY/MM/DD"""
    parts = re.split(r'[-/]', str(date_str))
    if len(parts) == 3:
        year, month, day = parts
        return f"{year}/{month.zfill(2)}/{day.zfill(2)}"
    return date_str


def parse_single_record(text: str) -> dict:
    """解析单条记录"""
    result = {"员工号": None, "姓名": None, "开始日期": None, "结束日期": None}
    emp = re.search(r'\b(\d{6})\b', text)
    if emp:
        result["员工号"] = emp.group(1)
    name = re.search(r'\d{6}\s*([\u4e00-\u9fa5]{2,4})', text)
    if name:
        result["姓名"] = name.group(1)
    dates = re.findall(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', text)
    if dates:
        result["开始日期"] = normalize_date(dates[0])
        result["结束日期"] = normalize_date(dates[1]) if len(dates) > 1 else normalize_date(dates[0])
    return result


def split_continuous_text(text: str) -> list:
    """把连续粘贴的文本按员工号切分成多条记录"""
    parts = re.split(r'(?=\d{6}[\u4e00-\u9fa5])', text)
    return [p.strip() for p in parts if p.strip() and re.search(r'\d{6}', p)]


def parse_batch_input(text: str, whitelist: set = None) -> tuple:
    """解析批量输入"""
    records = []
    errors = []
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])
    for i, line in enumerate(lines, 1):
        record = parse_single_record(line)
        if whitelist and record["员工号"] and record["员工号"] not in whitelist:
            continue
        if not record["员工号"]:
            errors.append(f"第{i}条: 未识别员工号 [{line[:50]}]")
            continue
        if not record["开始日期"]:
            errors.append(f"第{i}条: 未识别日期 [{line[:50]}]")
            continue
        records.append(record)
    return records, errors


def parse_excel_file(filepath: str, whitelist: set = None) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    if not HAS_OPENPYXL:
        return [], ["未安装openpyxl库，请运行: pip install openpyxl"]
    
    records = []
    errors = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            emp_id = str(int(row[0])).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            start_date = row[2] if len(row) > 2 else None
            end_date = row[3] if len(row) > 3 else None
            
            if not emp_id or not re.match(r'^\d{6}$', emp_id):
                if emp_id and emp_id != 'None':
                    errors.append(f"第{row_num}行: 员工号格式错误 [{emp_id}]")
                continue
            
            if whitelist and emp_id not in whitelist:
                continue
            
            def format_date(d):
                if isinstance(d, datetime):
                    return d.strftime('%Y/%m/%d')
                if isinstance(d, str):
                    return normalize_date(d)
                return None
            
            start = format_date(start_date)
            end = format_date(end_date) if end_date else start
            
            if not start:
                errors.append(f"第{row_num}行: 日期格式错误")
                continue
            
            records.append({
                "员工号": emp_id,
                "姓名": name,
                "开始日期": start,
                "结束日期": end or start
            })
        
        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")
    
    return records, errors


def fill_date(page, date_str):
    """填写日期到日期选择器"""
    parts = date_str.split('/')
    if len(parts) != 3:
        raise ValueError(f"日期格式错误: {date_str}")
    
    year, month, day = parts
    
    # 月份映射 - 注意11月和12月的特殊格式
    month_map = {
        '01': '一月', '02': '二月', '03': '三月', '04': '四月',
        '05': '五月', '06': '六月', '07': '七月', '08': '八月',
        '09': '九月', '10': '十月', '11': '十一', '12': '十二'
    }
    
    month_cn = month_map.get(month, month)
    day_str = str(int(day))  # 去掉前导0
    
    # 等待iframe加载
    page.wait_for_timeout(300)
    
    # 1. 点击年份输入框
    page.locator("iframe").nth(2).content_frame.get_by_role("textbox").nth(1).click()
    page.wait_for_timeout(100)
    
    # 2. 选择年份
    page.locator("iframe").nth(2).content_frame.get_by_role("cell", name=year).click()
    page.wait_for_timeout(100)
    
    # 3. 点击月份输入框
    page.locator("iframe").nth(2).content_frame.get_by_role("textbox").first.click()
    page.wait_for_timeout(100)
    
    # 4. 选择月份
    page.locator("iframe").nth(2).content_frame.get_by_role("cell", name=month_cn).click()
    page.wait_for_timeout(100)
    
    # 5. 选择日期 - 使用first选择第一个匹配的日期（避免匹配到15、25等）
    page.locator("iframe").nth(2).content_frame.get_by_role("cell", name=day_str).first.click()
    page.wait_for_timeout(100)


def query_flight_record(page, emp_id, start_date, end_date, clear_first=False):
    """执行查询操作
    
    Args:
        page: playwright页面对象
        emp_id: 员工号
        start_date: 开始日期 (格式: YYYY/MM/DD)
        end_date: 结束日期 (格式: YYYY/MM/DD)
        clear_first: 是否先清空表单（第一次查询不需要清空）
    """
    # 填写员工号
    emp_input = page.get_by_placeholder("员工号或姓名")
    emp_input.click()
    page.wait_for_timeout(100)
    
    if clear_first:
        # 第二次及以后：先清空再输入
        emp_input.fill("")
        page.wait_for_timeout(100)
    
    # 用type模拟逐字输入
    emp_input.type(str(emp_id), delay=50)
    page.wait_for_timeout(300)
    
    # 填写开始日期 - 多次点击确保触发日期选择器
    page.locator("#flyTimeExperience_beginDate").click()
    page.wait_for_timeout(200)
    if clear_first:
        # 第二次及以后需要多点几次
        page.locator("#flyTimeExperience_beginDate").click()
        page.wait_for_timeout(200)
    fill_date(page, start_date)
    
    # 填写结束日期 - 多次点击确保触发日期选择器
    page.locator("#flyTimeExperience_endDate").click()
    page.wait_for_timeout(200)
    if clear_first:
        # 第二次及以后需要多点几次
        page.locator("#flyTimeExperience_endDate").click()
        page.wait_for_timeout(200)
    fill_date(page, end_date)
    
    # 点击查询
    page.wait_for_timeout(300)
    page.get_by_role("button", name="查询").click()
    page.wait_for_timeout(1500)


def extract_flight_data(page):
    """提取飞行经历数据
    
    Returns:
        dict: 包含员工号、姓名、飞行经历、起落总数等信息
    """
    try:
        # 等待表格加载
        page.wait_for_timeout(1000)
        
        # 查找表格tbody
        tbody = page.locator("tbody.list")
        
        # 获取第一行数据（通常只有一条记录）
        first_row = tbody.locator("tr").first
        
        # 提取所有td单元格
        cells = first_row.locator("td")
        cell_count = cells.count()
        
        if cell_count == 0:
            return None
        
        # 按照表头顺序提取数据
        data = {
            "员工号": cells.nth(0).inner_text().strip() if cell_count > 0 else "",
            "姓名": cells.nth(1).inner_text().strip() if cell_count > 1 else "",
            "注册基地": cells.nth(2).inner_text().strip() if cell_count > 2 else "",
            "运行基地": cells.nth(3).inner_text().strip() if cell_count > 3 else "",
            "技术信息": cells.nth(4).inner_text().strip() if cell_count > 4 else "",
            "开始日期": cells.nth(5).inner_text().strip() if cell_count > 5 else "",
            "结束日期": cells.nth(6).inner_text().strip() if cell_count > 6 else "",
            "飞行时间": cells.nth(7).inner_text().strip() if cell_count > 7 else "",
            "飞行经历": cells.nth(8).inner_text().strip() if cell_count > 8 else "",
            "航段数": cells.nth(9).inner_text().strip() if cell_count > 9 else "",
            "夜航经历": cells.nth(10).inner_text().strip() if cell_count > 10 else "",
            "左座经历": cells.nth(11).inner_text().strip() if cell_count > 11 else "",
            "右座经历": cells.nth(12).inner_text().strip() if cell_count > 12 else "",
            "模拟机": cells.nth(13).inner_text().strip() if cell_count > 13 else "",
            "本场时间": cells.nth(14).inner_text().strip() if cell_count > 14 else "",
            "起落总数": cells.nth(15).inner_text().strip() if cell_count > 15 else "",
            "航线起落": cells.nth(16).inner_text().strip() if cell_count > 16 else "",
            "本场起落": cells.nth(17).inner_text().strip() if cell_count > 17 else "",
            "人工飞行时间": cells.nth(18).inner_text().strip() if cell_count > 18 else "",
        }
        
        return data
    except Exception as e:
        print(c_warn(f"提取数据失败: {e}"))
        return None


def whitelist_status(whitelist):
    """返回白名单状态文字"""
    if whitelist:
        return c_ok(f"白名单:{len(whitelist)}人")
    return c_warn("白名单:无")


def read_multiline(prompt, confirm_key='ok', cancel_key='c'):
    """读取多行输入"""
    print(prompt)
    lines = []
    while True:
        line = input()
        if line.lower() == cancel_key:
            return None
        if line.lower() == confirm_key:
            break
        if line:
            lines.append(line)
    if not lines:
        return None
    return '\n'.join(lines)


def set_whitelist():
    """设置白名单"""
    text = read_multiline(c_hint("请粘贴员工号列表(输入ok确认,c取消):"), 'ok', 'c')
    if text is None:
        print(c_warn("已取消"))
        return None
    wl = parse_whitelist(text)
    if not wl:
        print(c_err("未识别到有效员工号"))
        return None
    print(c_ok(f"已设置白名单,共{len(wl)}人"))
    return wl


def format_record(r):
    """格式化记录显示"""
    name = r.get('姓名') or '未知'
    return f"{r['员工号']} {name} {r['开始日期']}~{r['结束日期']}"


def print_failed_records(failed_records):
    """打印失败记录并写入日志文件"""
    if failed_records:
        print(c_err(f"本次失败{len(failed_records)}条:"))
        for r, reason in failed_records:
            print(c_err(f"  {format_record(r)} - {reason}"))
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"failed_{timestamp}.txt"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(f"失败记录 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"共{len(failed_records)}条\n")
                f.write("-" * 50 + "\n")
                for r, reason in failed_records:
                    f.write(f"{format_record(r)} - {reason}\n")
            filepath = os.path.abspath(filename)
            print(c_warn(f"失败记录已保存: {filepath}"))
        except Exception as e:
            print(c_warn(f"保存日志失败: {e}"))


def batch_mode(page, whitelist):
    """批量模式"""
    failed_records = []
    while True:
        print(f"{c_info('[批量模式]')} {whitelist_status(whitelist)}")
        text = read_multiline(c_hint("请粘贴数据(输入ok确认,b返回):"), 'ok', 'b')
        if text is None:
            print_failed_records(failed_records)
            return
        records, errors = parse_batch_input(text, whitelist)
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(err))
        if not records:
            print(c_err("没有可处理的记录"))
            continue
        print(c_ok(f"共{len(records)}条有效数据:"))
        for i, r in enumerate(records, 1):
            print(f"{i}. {format_record(r)}")
        confirm = input(c_hint("y开始查询,n重新粘贴,b返回主菜单: ")).strip().lower()
        if confirm == 'b':
            return
        if confirm != 'y':
            continue
        i = 0
        while i < len(records):
            record = records[i]
            print(f"{c_info(f'[{i+1}/{len(records)}]')} 查询: {format_record(record)}")
            try:
                # 第一条不清空表单，后续的清空
                query_flight_record(page, record["员工号"], record["开始日期"], record["结束日期"], clear_first=(i > 0))
                
                # 提取数据
                data = extract_flight_data(page)
                if data:
                    print(c_ok(f"查询完成 - 飞行经历: {data['飞行经历']} | 起落总数: {data['起落总数']}"))
                else:
                    print(c_ok("查询完成"))
                
                cmd = input(c_hint("回车继续查询,s跳过,b返回主菜单: ")).strip().lower()
                if cmd == 'b':
                    print_failed_records(failed_records)
                    return
                if cmd == 's':
                    failed_records.append((record, "手动跳过"))
                i += 1
            except Exception as e:
                beep_error()
                print(c_err(f"失败: {e}"))
                while True:
                    cmd = input(c_hint("s跳过,r重试,b返回主菜单: ")).strip().lower()
                    if cmd == 'b':
                        failed_records.append((record, str(e)))
                        print_failed_records(failed_records)
                        return
                    if cmd == 'r':
                        break
                    if cmd == 's':
                        failed_records.append((record, str(e)))
                        i += 1
                        break
                    print(c_warn("无效输入，请输入s/r/b"))
                if cmd == 'r':
                    continue
        print(c_ok("批量处理完成"))
        print_failed_records(failed_records)
        return


def manual_mode(page, whitelist):
    """手动模式"""
    query_count = 0  # 记录查询次数
    while True:
        print(f"{c_info('[手动模式]')} {whitelist_status(whitelist)} | {c_hint('粘贴数据,b返回主菜单:')}")
        text = input().strip()
        if text.lower() == 'b':
            return
        if not text:
            continue
        record = parse_single_record(text)
        if whitelist and record["员工号"] and record["员工号"] not in whitelist:
            print(c_err("该员工不在白名单中"))
            continue
        if not record["员工号"]:
            print(c_err("未识别员工号"))
            continue
        if not record["开始日期"]:
            print(c_err("未识别日期"))
            continue
        while True:
            print(f"查询: {format_record(record)}")
            try:
                # 第一次查询不清空，后续清空
                query_flight_record(page, record["员工号"], record["开始日期"], record["结束日期"], clear_first=(query_count > 0))
                query_count += 1
                
                # 提取数据
                data = extract_flight_data(page)
                if data:
                    print(c_ok(f"查询完成 - 飞行经历: {data['飞行经历']} | 起落总数: {data['起落总数']}"))
                else:
                    print(c_ok("查询完成"))
                break
            except Exception as e:
                beep_error()
                print(c_err(f"失败: {e}"))
                while True:
                    cmd = input(c_hint("r重试,b返回主菜单: ")).strip().lower()
                    if cmd == 'b':
                        return
                    if cmd == 'r':
                        break
                    print(c_warn("无效输入，请输入r/b"))
                if cmd == 'r':
                    continue
                break


def excel_mode(page, whitelist):
    """Excel导入模式"""
    if not HAS_OPENPYXL:
        print(c_err("未安装openpyxl库，请运行: pip install openpyxl"))
        return
    
    failed_records = []
    while True:
        print(f"{c_info('[Excel导入]')} {whitelist_status(whitelist)}")
        filepath = input(c_hint("请输入Excel文件路径(b返回): ")).strip()
        if filepath.lower() == 'b':
            print_failed_records(failed_records)
            return
        
        filepath = filepath.strip('"').strip("'")
        
        if not os.path.exists(filepath):
            print(c_err("文件不存在"))
            continue
        
        records, errors = parse_excel_file(filepath, whitelist)
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(f"  {err}"))
        
        if not records:
            print(c_err("没有可处理的记录"))
            continue
        
        print(c_ok(f"共{len(records)}条有效数据:"))
        for i, r in enumerate(records, 1):
            print(f"{i}. {format_record(r)}")
        
        confirm = input(c_hint("y开始查询,n重新选择,b返回主菜单: ")).strip().lower()
        if confirm == 'b':
            return
        if confirm != 'y':
            continue
        
        i = 0
        while i < len(records):
            record = records[i]
            print(f"{c_info(f'[{i+1}/{len(records)}]')} 查询: {format_record(record)}")
            try:
                # 第一条不清空表单，后续的清空
                query_flight_record(page, record["员工号"], record["开始日期"], record["结束日期"], clear_first=(i > 0))
                
                # 提取数据
                data = extract_flight_data(page)
                if data:
                    print(c_ok(f"查询完成 - 飞行经历: {data['飞行经历']} | 起落总数: {data['起落总数']}"))
                else:
                    print(c_ok("查询完成"))
                
                cmd = input(c_hint("回车继续查询,s跳过,b返回主菜单: ")).strip().lower()
                if cmd == 'b':
                    print_failed_records(failed_records)
                    return
                if cmd == 's':
                    failed_records.append((record, "手动跳过"))
                i += 1
            except Exception as e:
                beep_error()
                print(c_err(f"失败: {e}"))
                while True:
                    cmd = input(c_hint("s跳过,r重试,b返回主菜单: ")).strip().lower()
                    if cmd == 'b':
                        failed_records.append((record, str(e)))
                        print_failed_records(failed_records)
                        return
                    if cmd == 'r':
                        break
                    if cmd == 's':
                        failed_records.append((record, str(e)))
                        i += 1
                        break
                    print(c_warn("无效输入，请输入s/r/b"))
                if cmd == 'r':
                    continue
        print(c_ok("Excel导入完成"))
        print_failed_records(failed_records)
        return


def main():
    print(c_info("飞行经历起落数查询助手"))
    
    browser_path = input(c_hint("浏览器路径(回车用默认): ")).strip() or None
    if browser_path:
        print(c_ok(f"使用指定浏览器: {browser_path}"))
    else:
        print(c_ok("使用默认浏览器"))
    
    whitelist = None
    use_wl = input(c_hint("是否预设白名单?(y/n): ")).strip().lower()
    if use_wl == 'y':
        whitelist = set_whitelist()
    else:
        print(c_ok("不设置白名单,处理所有员工"))
    
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=False, executable_path=browser_path)
    context = browser.new_context()
    context.set_default_timeout(0)
    page = context.new_page()
    
    try:
        page.goto("https://ieb.csair.com/login")
        page.wait_for_load_state("networkidle")
        page.locator("#scanLogin").wait_for()
        page.locator("#scanLogin").click()
        print(c_info("请扫码登录..."))
        page.wait_for_url("**/index/**")
        page.wait_for_load_state("networkidle")
        print(c_ok("登录成功"))
    except Exception as e:
        print(c_err(f"自动登录失败: {e}"))
        print(c_warn("请手动完成登录"))
        input(c_hint("登录完成后按回车继续..."))
    
    try:
        print(c_info("正在进入飞行经历查询页面..."))
        page.goto("https://ieb.csair.com/index/index")
        page.wait_for_load_state("networkidle")
        page.get_by_text("统计应用").nth(1).wait_for()
        page.get_by_text("统计应用").nth(1).click()
        page.get_by_role("link", name="综合报表").wait_for()
        page.get_by_role("link", name="综合报表").click()
        page.get_by_role("link", name="飞行经历").wait_for()
        page.get_by_role("link", name="飞行经历").click()
        page.wait_for_load_state("networkidle")
        # 进入页面后立即选择"按员工号查询"单选按钮
        page.wait_for_timeout(500)
        page.get_by_role("radio").nth(2).check()
        page.wait_for_timeout(300)
        print(c_ok("已进入飞行经历查询页面"))
    except Exception as e:
        print(c_err(f"自动导航失败: {e}"))
        print(c_warn("请手动进入飞行经历查询页面"))
        input(c_hint("准备好后按回车继续..."))
    
    print(c_ok("开始工作"))
    while True:
        print(f"{whitelist_status(whitelist)} | {c_hint('1批量 2手动 3Excel导入 w设白名单 c清白名单 q退出')}")
        cmd = input(c_hint("选择: ")).strip().lower()
        if cmd == '1':
            batch_mode(page, whitelist)
        elif cmd == '2':
            manual_mode(page, whitelist)
        elif cmd == '3':
            excel_mode(page, whitelist)
        elif cmd == 'w':
            new_wl = set_whitelist()
            if new_wl is not None:
                whitelist = new_wl
        elif cmd == 'c':
            whitelist = None
            print(c_ok("已清除白名单"))
        elif cmd == 'q':
            break
    
    print(c_info("\n程序结束，浏览器保持打开状态"))
    print(c_hint("按回车关闭程序..."))
    input()


if __name__ == "__main__":
    main()
